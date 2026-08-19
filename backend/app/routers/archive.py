"""网页任务归档：一键导出 ZIP（Excel 汇总 + 按人分类资料）与导入还原。

ZIP 目录结构（模板化，人工可读、机器可还原）：
  <任务名>_<导出时间>/
    manifest.json                      — 完整状态（卡片/进度/报告，不含图片二进制）
    汇总表.xlsx                        — Sheet1 汇总（一人一行），Sheet2 字段对比明细
    学生资料/
      01_张三_<rid>/
        资料信息.json                  — 该人字段、护照字段、提取字段、报告摘要
        卡片图片.jpg                   — 裁剪好的卡片图
        头像.jpg                       — 头像
        提取文件/
          passport.jpg                 — 预处理后的提取文件图片（裁剪/转正后）
          passport_全文.txt            — 提取全文

导入时按 manifest + 文件引用还原全部状态（图片重新内联为 base64 返回前端）。
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from datetime import datetime
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api/archive", tags=["archive"])

OVERALL_LABELS = {"pass": "通过", "fail": "有问题", "review": "需检查"}
MATCH_LABELS = {
    "match": "一致", "mismatch": "不一致", "missing": "缺失",
    "partial": "部分", "unknown": "未知", "error": "错误",
}


def _safe_name(name: str, fallback: str) -> str:
    """文件夹/文件名净化：去非法字符与空白，限制长度。"""
    s = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", (name or "").strip())
    s = re.sub(r"_+", "_", s).strip("._ ")
    return (s[:40] or fallback)


def _b64_bytes(b64: str) -> bytes:
    import base64
    return base64.b64decode(b64)


def _bytes_b64(data: bytes) -> str:
    import base64
    return base64.b64encode(data).decode("ascii")


@router.post("/export")
async def export_archive(payload: dict):
    """前端收集的全部任务状态 → 打包 ZIP 下载。"""
    records = payload.get("records") or []
    if not records:
        raise HTTPException(400, "没有可导出的任务卡片")

    task_name = _safe_name(payload.get("task_name") or "网页任务", "网页任务")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    root = f"{task_name}_{stamp}"
    reports = payload.get("reports") or []

    # 报告按 record_id 索引（汇总表与资料信息用）
    reports_by_record: dict[str, dict] = {}
    for rp in reports:
        if isinstance(rp, dict) and rp.get("record_id"):
            reports_by_record[str(rp["record_id"])] = rp

    buf = io.BytesIO()
    manifest_records: list[dict] = []

    # ---- 收集全部字段名（Excel 动态列）----
    field_keys: list[str] = []
    for r in records:
        for k in (r.get("fields") or {}):
            if k not in field_keys:
                field_keys.append(k)

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, r in enumerate(records, 1):
            rid = str(r.get("record_id"))
            disp = _safe_name(str(r.get("display_name") or rid), rid)
            folder = f"{root}/学生资料/{idx:02d}_{disp}_{rid[:8]}"

            # manifest 记录：不含图片二进制，只含文件引用
            m_rec: dict[str, Any] = {
                "record_id": rid,
                "source": r.get("source"),
                "display_name": r.get("display_name") or "",
                "fields": r.get("fields") or {},
                "university_url": r.get("university_url"),
                "university_name": r.get("university_name"),
                "has_passport": r.get("has_passport"),
                "passport_fields": r.get("passport_fields") or {},
                "overall": r.get("overall"),
                "loop": r.get("loop"),
            }
            avatar_b64 = (r.get("avatar_b64") or "").strip()
            if avatar_b64:
                zf.writestr(f"{folder}/头像.jpg", _b64_bytes(avatar_b64))
                m_rec["avatar_file"] = "头像.jpg"
            card_b64 = (r.get("card_image_b64") or "").strip()
            if card_b64:
                zf.writestr(f"{folder}/卡片图片.jpg", _b64_bytes(card_b64))
                m_rec["card_image_file"] = "卡片图片.jpg"

            # 提取文件：图片 + 全文
            docs_m: list[dict] = []
            for di, d in enumerate(r.get("docs") or [], 1):
                base = _safe_name(str(d.get("filename") or f"文件{di}"), f"文件{di}")
                d_m = {
                    "filename": d.get("filename") or base,
                    "method": d.get("method") or "",
                    "ocr_backend": d.get("ocr_backend") or "",
                    "fields": d.get("fields") or {},
                    "source": d.get("source") or "",
                    "mrz_warnings": d.get("mrz_warnings") or [],
                }
                img_b64 = (d.get("image_b64") or "").strip()
                if img_b64:
                    img_name = f"{base}.jpg"
                    zf.writestr(f"{folder}/提取文件/{img_name}", _b64_bytes(img_b64))
                    d_m["image_file"] = f"提取文件/{img_name}"
                text = str(d.get("text") or "")
                if text.strip():
                    txt_name = f"{base}_全文.txt"
                    zf.writestr(f"{folder}/提取文件/{txt_name}", text.encode("utf-8"))
                    d_m["text_file"] = f"提取文件/{txt_name}"
                docs_m.append(d_m)
            m_rec["docs"] = docs_m

            # 资料信息.json：给人看的分类资料（含报告摘要）
            rp = reports_by_record.get(rid)
            person_info = {
                "姓名": m_rec["display_name"],
                "状态": OVERALL_LABELS.get(m_rec["overall"], "未执行") if m_rec["overall"] else "未执行",
                "LOOP模板": (m_rec["loop"] or {}).get("loopName", "") if m_rec["loop"] else "",
                "学校": m_rec["university_name"] or "",
                "字段": m_rec["fields"],
                "护照字段": m_rec["passport_fields"],
                "提取文件": [
                    {"文件": d["filename"], "方式": d["method"], "字段": d["fields"],
                     **({"MRZ警告": d["mrz_warnings"]} if d["mrz_warnings"] else {})}
                    for d in docs_m
                ],
            }
            if rp:
                bad = [e for e in (rp.get("entries") or []) if e.get("match") not in ("match", None)]
                person_info["审查报告"] = {
                    "结论": OVERALL_LABELS.get(rp.get("overall"), rp.get("overall") or ""),
                    "问题字段": [
                        {"字段": e.get("right_label") or e.get("left_field") or "",
                         "网页值": e.get("right_value"), "期望值": e.get("left_value"),
                         "结果": MATCH_LABELS.get(e.get("match"), str(e.get("match")))}
                        for e in bad
                    ],
                    "开始时间": rp.get("started_at"),
                    "结束时间": rp.get("finished_at"),
                    "摘要": rp.get("summary") or "",
                }
            zf.writestr(
                f"{folder}/资料信息.json",
                json.dumps(person_info, ensure_ascii=False, indent=2),
            )
            manifest_records.append(m_rec)

        # ---- manifest.json：完整状态（可整体还原）----
        manifest = {
            "version": 1,
            "app": "CINSIDE",
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "task_name": payload.get("task_name") or "网页任务",
            "run_cursor": payload.get("run_cursor"),
            "records": manifest_records,
            # 报告剥离截图（体积大），保留全部比对数据
            "reports": [
                {
                    **{k: v for k, v in rp.items() if k != "entries"},
                    "entries": [
                        {k: v for k, v in e.items() if k != "screenshot"}
                        for e in (rp.get("entries") or [])
                    ],
                }
                for rp in reports if isinstance(rp, dict)
            ],
        }
        zf.writestr(f"{root}/manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

        # ---- 汇总表.xlsx ----
        zf.writestr(f"{root}/汇总表.xlsx", _build_excel(manifest_records, reports_by_record, field_keys))

    buf.seek(0)
    filename = f"{root}.zip"
    # 文件名含中文：HTTP 头只允许 latin-1，改用百分号编码（前端 decodeURIComponent 还原）
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{encoded}\"; filename*=UTF-8''{encoded}"
            )
        },
    )


def _build_excel(records: list[dict], reports_by_record: dict, field_keys: list[str]) -> bytes:
    """汇总表：Sheet1 一人一行汇总；Sheet2 字段对比明细（来自报告 entries）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    head_font = Font(bold=True, size=10)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")

    status_color = {"通过": "C6EFCE", "有问题": "FFC7CE", "需检查": "FFEB9C"}

    headers = ["序号", "姓名", "状态", "LOOP模板", "学校", *field_keys, "提取文件数", "问题字段数"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = head_font
        c.alignment = center

    for i, r in enumerate(records, 1):
        rp = reports_by_record.get(r["record_id"])
        bad_n = sum(1 for e in (rp.get("entries") or []) if e.get("match") not in ("match", None)) if rp else 0
        overall = OVERALL_LABELS.get(r.get("overall"), "未执行") if r.get("overall") else "未执行"
        row = [
            i,
            r.get("display_name") or r["record_id"],
            overall,
            (r.get("loop") or {}).get("loopName", "") if r.get("loop") else "",
            r.get("university_name") or "",
            *[ (r.get("fields") or {}).get(k, "") for k in field_keys ],
            len(r.get("docs") or []),
            bad_n,
        ]
        ws.append(row)
        # 状态列着色
        cell = ws.cell(row=ws.max_row, column=3)
        cell.fill = PatternFill("solid", fgColor=status_color.get(overall, "F3F4F6"))

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    for j in range(6, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 14
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body_font

    # Sheet2：字段对比明细
    ws2 = wb.create_sheet("字段对比明细")
    ws2.append(["序号", "姓名", "字段", "网页值", "期望值", "结果", "说明"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = head_font
        c.alignment = center
    n = 0
    for r in records:
        rp = reports_by_record.get(r["record_id"])
        if not rp:
            continue
        for e in rp.get("entries") or []:
            n += 1
            ws2.append([
                n,
                r.get("display_name") or r["record_id"],
                e.get("right_label") or e.get("left_field") or "",
                str(e.get("right_value") or ""),
                str(e.get("left_value") or ""),
                MATCH_LABELS.get(e.get("match"), str(e.get("match"))),
                e.get("reasoning") or "",
            ])
            mc = ws2.cell(row=ws2.max_row, column=6)
            mc.fill = PatternFill(
                "solid",
                fgColor="C6EFCE" if e.get("match") == "match" else "FFC7CE",
            )
    for col, w in zip("ABCDEFG", (6, 14, 16, 22, 22, 9, 24)):
        ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.font = body_font

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.post("/import")
async def import_archive(file: UploadFile = File(...)):
    """上传归档 ZIP → 解析 manifest + 文件引用 → 还原 payload（图片内联 base64）。

    返回与导出时前端 payload 同构的数据，前端直接恢复各状态。
    """
    data = await file.read()
    if not data:
        raise HTTPException(400, "空文件")
    if len(data) > 500 * 1024 * 1024:
        raise HTTPException(400, "文件过大（>500MB）")

    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception:
        raise HTTPException(400, "不是有效的 ZIP 归档")

    # 定位 manifest.json（兼容外层多包了一层目录的情况）
    manifest_path = next(
        (n for n in zf.namelist() if n.rstrip("/").endswith("manifest.json")), None
    )
    if not manifest_path:
        raise HTTPException(400, "归档缺少 manifest.json（可能不是 CINSIDE 任务归档）")
    root_prefix = manifest_path[: -len("manifest.json")]

    try:
        manifest = json.loads(zf.read(manifest_path).decode("utf-8"))
    except Exception:
        raise HTTPException(400, "manifest.json 解析失败")

    def read_ref(base_dir: str, ref: str) -> str:
        """读归档内文件 → base64。"""
        try:
            return _bytes_b64(zf.read(f"{base_dir}/{ref}"))
        except Exception:
            return ""

    # 人物资料目录集合：学生资料/<序号_姓名_rid前缀>（namelist 只含文件，目录需从路径提取）
    stu_prefix = root_prefix + "学生资料/"
    person_dirs = {
        n[len(stu_prefix):].split("/")[0]
        for n in zf.namelist()
        if n.startswith(stu_prefix) and "/" in n[len(stu_prefix):]
    }

    records_out = []
    for r in manifest.get("records") or []:
        rid = str(r.get("record_id"))
        rec = {
            "record_id": rid,
            "source": r.get("source") or "manual",
            "display_name": r.get("display_name") or "",
            "fields": r.get("fields") or {},
            "university_url": r.get("university_url"),
            "university_name": r.get("university_name"),
            "has_passport": bool(r.get("has_passport")),
            "passport_fields": r.get("passport_fields") or {},
            "overall": r.get("overall"),
            "loop": r.get("loop"),
            "avatar_b64": "",
            "card_image_b64": "",
            "docs": [],
        }
        # 定位该记录的资料目录（目录名含 record_id 前 8 位）
        dir_name = next(
            (d for d in person_dirs if rid[:8] and rid[:8] in d),
            "",
        )
        folder_dir = f"{stu_prefix}{dir_name}" if dir_name else ""
        if folder_dir:
            if r.get("avatar_file"):
                rec["avatar_b64"] = read_ref(folder_dir, r["avatar_file"])
            if r.get("card_image_file"):
                rec["card_image_b64"] = read_ref(folder_dir, r["card_image_file"])
        for d in r.get("docs") or []:
            doc = {
                "filename": d.get("filename") or "文件",
                "method": d.get("method") or "",
                "ocr_backend": d.get("ocr_backend") or "",
                "fields": d.get("fields") or {},
                "source": d.get("source") or "",
                "mrz_warnings": d.get("mrz_warnings") or [],
                "text": "",
                "image_b64": "",
            }
            if folder_dir:
                if d.get("text_file"):
                    try:
                        doc["text"] = zf.read(f"{folder_dir}/{d['text_file']}").decode("utf-8", errors="replace")
                    except Exception:
                        pass
                if d.get("image_file"):
                    doc["image_b64"] = read_ref(folder_dir, d["image_file"])
            rec["docs"].append(doc)
        records_out.append(rec)

    return {
        "version": manifest.get("version"),
        "exported_at": manifest.get("exported_at"),
        "task_name": manifest.get("task_name") or "网页任务",
        "run_cursor": manifest.get("run_cursor"),
        "records": records_out,
        "reports": manifest.get("reports") or [],
    }
