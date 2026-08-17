"""上传与数据源路由。"""
from __future__ import annotations

import os
import uuid
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Response, UploadFile
from pydantic import BaseModel

from ..config import settings
from ..models import ApplicantRecord, PassportData
from ..services.excel_parser import parse_bytes
from ..services.passport_ocr import extract_passport
from ..services.task_manager import clear_right_records, list_records, list_right_records, upsert_passport, upsert_records, upsert_right_records

router = APIRouter(prefix="/api", tags=["upload"])

# 上传源信息：记录原始文件路径/字节，用于「导出 Excel」时把内存中的修正值写回原文件
# kind: path=本地文件（Electron 直接传路径，可原地写回）；bytes=浏览器上传（仅内存，导出为副本）
LEFT_SOURCE: dict = {}
RIGHT_SOURCE: dict = {}


@router.post("/upload/excel")
async def upload_excel(file: UploadFile = File(...)):
    """上传 Excel/CSV，解析成 records（左侧数据源）。"""
    content = await file.read()
    if not content:
        raise HTTPException(400, "empty file")
    try:
        records, detected_map = await parse_bytes(content, file.filename or "upload.xlsx")
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    except Exception as e:
        raise HTTPException(400, f"解析失败: {e}")

    upsert_records(records)
    # 保存上传源字节：导出时可在原始布局上写回修正值（Electron 本地路径会再覆盖为 path 模式）
    LEFT_SOURCE.clear()
    LEFT_SOURCE.update({"kind": "bytes", "bytes": content, "filename": file.filename or "data.xlsx"})
    return {
        "count": len(records),
        "records": [r.model_dump() for r in records],
        "detected_column_map": detected_map,
    }


@router.post("/upload/excel-right")
async def upload_excel_right(file: UploadFile = File(...)):
    """上传右侧参考 Excel/CSV，解析成 right_records（不覆盖左侧数据源）。"""
    content = await file.read()
    if not content:
        raise HTTPException(400, "empty file")
    try:
        records, detected_map = await parse_bytes(content, file.filename or "upload.xlsx")
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    except Exception as e:
        raise HTTPException(400, f"解析失败: {e}")

    upsert_right_records(records)
    RIGHT_SOURCE.clear()
    RIGHT_SOURCE.update({"kind": "bytes", "bytes": content, "filename": file.filename or "data.xlsx"})
    return {
        "count": len(records),
        "records": [r.model_dump() for r in records],
        "detected_column_map": detected_map,
    }


class ExcelSourceBody(BaseModel):
    """Electron 渲染层可拿到 File.path：登记本地路径后导出可直接原地写回原文件。"""
    side: str = "left"
    path: str = ""
    filename: str = ""


@router.post("/upload/excel-source")
def set_excel_source(body: ExcelSourceBody):
    target = LEFT_SOURCE if body.side == "left" else RIGHT_SOURCE
    if body.path and os.path.exists(body.path):
        target.clear()
        target.update({"kind": "path", "path": body.path, "filename": body.filename or os.path.basename(body.path)})
    elif body.filename:
        target["filename"] = body.filename
    return {"ok": True, "mode": target.get("kind", "bytes")}


def _apply_records_to_first_sheet(wb, recs, highlights: Optional[dict] = None) -> int:
    """把内存 records 的字段值写回工作簿第一个有数据的 sheet（按 rec-XXX 行序对应）。

    highlights: {record_id: {字段名: match状态}} —— 有问题的单元格填色高亮：
    mismatch/error=红色填充（红底深红字），missing/partial=琥珀填充（跟字段对比卡片色系一致）。
    """
    import logging
    _export_log = logging.getLogger("cinside.export")

    ws = None
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(50, sheet.max_row or 1)):
            if any(c.value is not None and str(c.value).strip() != "" for c in row):
                ws = sheet
                break
        if ws:
            break
    if ws is None:
        return 0
    # 表头行：第一个非空行；记录 列名 -> 列号
    headers: dict[str, int] = {}
    header_row = 0
    for row in ws.iter_rows(min_row=1):
        if any(c.value is not None and str(c.value).strip() != "" for c in row):
            header_row = row[0].row
            for c in row:
                if c.value is not None and str(c.value).strip():
                    headers[str(c.value).strip()] = c.column
            break
    if not header_row:
        return 0

    from openpyxl.styles import Font, PatternFill

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)
    amber_fill = PatternFill("solid", fgColor="FFEB9C")
    amber_font = Font(color="9C6500")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100", bold=True)
    # 用户已修复（含跨列连带修复）：亮黄填充，与琥珀（缺失/部分）区分
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    yellow_font = Font(color="7F6000", bold=True)

    def _hl_cell(cell, status: str) -> None:
        """按对比状态给单元格上高亮：红=不一致/错误，琥珀=缺失/部分，亮黄=用户已修复（值被改），绿=人工确认一致（值未动）。"""
        s = (status or "").lower()
        if s in ("mismatch", "error"):
            cell.fill = red_fill
            cell.font = red_font
        elif s in ("missing", "partial"):
            cell.fill = amber_fill
            cell.font = amber_font
        elif s in ("fixed",):
            cell.fill = yellow_fill
            cell.font = yellow_font
        elif s in ("match", "confirmed"):
            cell.fill = green_fill
            cell.font = green_font

    highlights = highlights or {}
    # 高亮键兼容：报告的 left_field 是规范字段名（surname/passport_no），
    # 而表头是原始列名（姓/护照号/姓 Family Name）——直接 key in rec_hl 永远对不上。
    # 分词精确匹配（避免子串误匹配：如 "name" 不应匹配 "family name"）
    from ..services.excel_parser import _normalize_key, _FIELD_ALIASES
    import re as _re

    def _header_to_canonical(header_key: str) -> list[str]:
        """把 Excel 表头转成所有可能的规范字段名列表。"""
        candidates: list[str] = []
        hk = _normalize_key(header_key)
        candidates.append(hk)
        hk_lower = hk.strip().lower()

        # 1. 完整字符串精确匹配别名表
        for alias, canonical in _FIELD_ALIASES.items():
            if alias.strip().lower() == hk_lower:
                candidates.append(canonical)
                break

        # 2. 按空格/分隔符分词，去除标点，逐词匹配别名表
        raw_tokens = _re.split(r'[\s,/\-]+', hk_lower)
        tokens = [_re.sub(r'[^\w\u4e00-\u9fff]', '', t) for t in raw_tokens if t.strip()]

        # 2a. 连续词组（2-3 词）优先匹配（处理 "Family Name"→surname, "Passport No."→passport_no）
        for i in range(len(tokens)):
            for j in range(i + 2, min(i + 4, len(tokens) + 1)):
                phrase = ' '.join(tokens[i:j])
                for alias, canonical in _FIELD_ALIASES.items():
                    if alias.strip().lower() == phrase:
                        candidates.append(canonical)
                        break

        # 2b. 单词精确匹配（处理 "姓"→surname, "国籍"→nationality, "护照号"→passport_no）
        for t in tokens:
            for alias, canonical in _FIELD_ALIASES.items():
                if alias.strip().lower() == t:
                    candidates.append(canonical)
                    break

        return candidates

    def _hl_lookup(rec_hl: dict, header_key: str) -> Optional[str]:
        st = rec_hl.get(header_key)
        if st:
            return st
        for ck in _header_to_canonical(header_key):
            st = rec_hl.get(ck)
            if st:
                return st
        return None

    written = 0
    data_idx = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        if all(c.value is None or str(c.value).strip() == "" for c in row):
            continue  # 解析时同样跳过空行，行序保持一致
        data_idx += 1
        rec = next((r for r in recs if r.record_id == f"rec-{data_idx:03d}"), None)
        if not rec:
            continue
        rec_hl = highlights.get(rec.record_id, {})
        # 高亮键先归一化展开（原始名 + 规范名都能命中）
        norm_hl: dict[str, str] = {}
        for k, st in rec_hl.items():
            kk = str(k).strip()
            norm_hl.setdefault(kk, st)
            norm_hl.setdefault(_normalize_key(kk), st)
        rec_hl = norm_hl
        for key, col in headers.items():
            status = _hl_lookup(rec_hl, key)
            if key in rec.fields:
                cell = ws.cell(row=row[0].row, column=col, value=rec.fields[key])
            elif status:
                # 字段在内存里没有值（如缺失）：不写值，仅上色标记问题格
                cell = ws.cell(row=row[0].row, column=col)
            else:
                continue
            if status:
                _hl_cell(cell, status)
        written += 1
    return written


@router.get("/upload/excel-export")
def export_excel(side: str = "left"):
    """导出 Excel（无高亮）：把内存中的（修正后）字段值写回文件。"""
    return _do_export(side, None)


class RecordHighlight(BaseModel):
    """单条记录的字段对比状态：{字段名: match/mismatch/missing/error/partial}。"""
    record_id: str
    fields: dict[str, str] = {}


class ExcelHighlightBody(BaseModel):
    """带高亮导出请求体：运行结束后前端把字段对比结果一起传上来。"""
    side: str = "left"
    highlights: list[RecordHighlight] = []


@router.post("/upload/excel-export-hl")
def export_excel_highlighted(body: ExcelHighlightBody):
    """导出 Excel（带问题高亮）：写回修正值的同时，把有问题的单元格填色。
    - mismatch/error → 红色填充+深红加粗字（同字段对比卡片"需检查"语义）
    - missing/partial → 琥珀填充（同"缺失"语义）
    """
    hl_map = {h.record_id: h.fields for h in body.highlights if h.record_id}
    return _do_export(body.side or "left", hl_map or None)


def _downloads_dir(fallback_dir: str) -> str:
    """系统「下载」文件夹；取不到或不存在时回退到原文件所在目录。"""
    home = os.path.expanduser("~")
    dl = os.path.join(home, "Downloads")
    if os.path.isdir(dl):
        return dl
    return os.path.dirname(os.path.abspath(fallback_dir)) or home


def _move_with_retry(src: str, dst: str, attempts: int = 3, delay: float = 0.25) -> None:
    """os.replace 带重试：OneDrive/Defender 对新文件有瞬时限权，稍候即恢复；
    目标被 Excel/WPS 长期占用时重试耗尽仍抛 PermissionError（由调用方兜底）。"""
    import time
    last: PermissionError | None = None
    for _ in range(attempts):
        try:
            os.replace(src, dst)
            return
        except PermissionError as e:
            last = e
            time.sleep(delay)
    raise last  # type: ignore[misc]


def _deliver_export(tmp_path: str, dl_dir: str, stem: str, ext: str, fallback_dir: str) -> tuple[str, str]:
    """把导出结果投递到「下载」文件夹，四级兜底保证导出永不失败：
    1. <原名>_审核结果.xlsx（常规，覆盖上一份）
    2. 同名+时间戳（上一份正被 Excel/WPS 打开）
    3. 复制写入（move 被同步软件/策略拦截但创建允许时）
    4. 回退原文件所在目录（下载目录整体不可写时）
    返回 (最终路径, 提示note)；全部失败抛 400。
    """
    import datetime
    import shutil

    target = os.path.join(dl_dir, f"{stem}_审核结果{ext}")
    try:
        _move_with_retry(tmp_path, target)
        return target, ""
    except PermissionError:
        pass
    ts = datetime.datetime.now().strftime("%H%M%S")
    alt = os.path.join(dl_dir, f"{stem}_审核结果_{ts}{ext}")
    try:
        _move_with_retry(tmp_path, alt)
        return alt, f"上一份导出结果正被 Excel/WPS 打开，已带时间戳另存：{alt}"
    except PermissionError:
        pass
    try:
        shutil.copyfile(tmp_path, target)
        return target, ""
    except PermissionError:
        pass
    fb = os.path.join(fallback_dir, f"{stem}_审核结果{ext}")
    try:
        _move_with_retry(tmp_path, fb, attempts=2)
        return fb, f"下载文件夹暂时无法写入，已导出到原文件旁边：{fb}"
    except PermissionError:
        raise HTTPException(400, f"导出文件无法写入（下载文件夹与原目录均被占用或无权限）：\n{target}\n请关闭 Excel/WPS 后重试")


def _do_export(side: str, highlights: Optional[dict]):
    """导出 Excel：把内存中的（修正后）字段值写成结果文件。
    - 本地路径模式：在原文件布局基础上，导出到「下载」文件夹 <原名>_审核结果.xlsx，
      原文件本身不动（可能藏在微信/OneDrive 深层目录，写回用户找不到）
    - 字节模式：在原始布局上写回后作为附件下载
    """
    import io

    from ..services.task_manager import store
    src = LEFT_SOURCE if side == "left" else RIGHT_SOURCE
    recs = list(store.records.values()) if side == "left" else list(store.right_records.values())
    if not recs:
        raise HTTPException(400, "没有可导出的数据，请先上传 Excel")

    path = src.get("path")
    if path and os.path.exists(path) and str(path).lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        _apply_records_to_first_sheet(wb, recs, highlights)
        # 导出统一落「下载」文件夹：<原名>_审核结果.xlsx——原文件可能藏在
        # 微信/OneDrive 等深层目录，原地写回用户根本找不到；原文件保持不动。
        stem, ext = os.path.splitext(os.path.basename(path))
        # 临时文件写系统 temp（下载文件夹可能被 OneDrive/Defender 盯梢，
        # 新建 zip 会被瞬时限权），写完由 _deliver_export 投递+兜底
        import tempfile
        fd, tmp_path = tempfile.mkstemp(suffix=ext, prefix="cinside_export_")
        os.close(fd)
        try:
            wb.save(tmp_path)
            target, note = _deliver_export(tmp_path, _downloads_dir(path), stem, ext, os.path.dirname(os.path.abspath(path)))
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
        resp = {"ok": True, "mode": "inplace", "path": target, "count": len(recs)}
        if note:
            resp["note"] = note
        return resp

    data = src.get("bytes")
    if data and (src.get("filename") or "").lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        _apply_records_to_first_sheet(wb, recs, highlights)
        buf = io.BytesIO()
        wb.save(buf)
        out_name = os.path.splitext(src.get("filename") or "data.xlsx")[0] + "_updated.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={out_name}"},
        )

    # 兜底：无原始文件（如 CSV 上传）→ 由内存数据新建 xlsx（同样带高亮）
    from openpyxl import Workbook
    cols: list[str] = []
    for r in recs:
        for k in r.fields:
            if not k.startswith("_") and k not in cols:
                cols.append(k)
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(cols)
    from openpyxl.styles import Font, PatternFill
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)
    amber_fill = PatternFill("solid", fgColor="FFEB9C")
    amber_font = Font(color="9C6500")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    yellow_font = Font(color="7F6000", bold=True)
    for r in recs:
        ws.append([r.fields.get(c, "") for c in cols])
        rec_hl = (highlights or {}).get(r.record_id, {})
        for ci, col in enumerate(cols, start=1):
            if col in rec_hl:
                s = (rec_hl[col] or "").lower()
                cell = ws.cell(row=ws.max_row, column=ci)
                if s in ("mismatch", "error"):
                    cell.fill = red_fill
                    cell.font = red_font
                elif s in ("missing", "partial"):
                    cell.fill = amber_fill
                    cell.font = amber_font
                elif s in ("fixed",):
                    cell.fill = yellow_fill
                    cell.font = yellow_font
    buf = io.BytesIO()
    wb.save(buf)
    out_name = os.path.splitext(src.get("filename") or "data.xlsx")[0] + "_updated.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={out_name}"},
    )


@router.post("/upload/passport/{record_id}")
async def upload_passport(record_id: str, file: UploadFile = File(...)):
    """为指定 record 上传一张护照图片，触发 Vision LLM 抽取。"""
    content = await file.read()
    if not content:
        raise HTTPException(400, "empty file")

    os.makedirs(settings.upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or ".jpg")[1] or ".jpg"
    save_name = f"{record_id}_{uuid.uuid4().hex[:8]}{ext}"
    save_path = os.path.join(settings.upload_dir, save_name)
    with open(save_path, "wb") as f:
        f.write(content)

    p: PassportData = await extract_passport(save_path, record_id, file.filename or save_name)
    upsert_passport(record_id, p)
    return p.model_dump()


@router.get("/records")
def get_records():
    """列出当前所有记录（含已上传护照数据）。"""
    from ..services.task_manager import get_passport
    out = []
    for r in list_records():
        d = r.model_dump()
        p = get_passport(r.record_id)
        d["has_passport"] = p is not None
        d["passport_fields"] = p.fields if p else {}
        out.append(d)
    return {"records": out}


@router.get("/records-right")
def get_right_records():
    """列出右侧参考Excel的所有记录。"""
    return {"records": [r.model_dump() for r in list_right_records()]}


@router.delete("/records-right")
def clear_right_records_endpoint():
    """清空右侧参考Excel数据。"""
    clear_right_records()
    return {"ok": True}


@router.delete("/records")
def clear_records():
    """清空所有数据（演示用）。"""
    from ..services.task_manager import store
    store.records.clear()
    store.passports.clear()
    return {"ok": True}


class FieldsUpdate(BaseModel):
    """更新记录字段请求体（审查修正：以来源值为准修正 Excel 字段）。"""
    fields: dict[str, str]


@router.patch("/records/{record_id}/fields")
def update_record_fields(record_id: str, body: FieldsUpdate):
    """更新指定记录的一个或多个字段值。"""
    from ..services.task_manager import store
    rec = store.records.get(record_id)
    if not rec:
        raise HTTPException(404, f"record {record_id} not found")
    for k, v in body.fields.items():
        rec.fields[k] = v
    return {"ok": True, "record_id": record_id, "updated": list(body.fields.keys())}


class AvatarUpdate(BaseModel):
    """更新头像请求体。"""
    avatar: str  # base64 字符串（无 data: 前缀）或图片 URL


@router.patch("/records/{record_id}/avatar")
def update_avatar(record_id: str, body: AvatarUpdate):
    """更新指定记录的头像。

    avatar 可以是：
    - base64 字符串（无 data:image/... 前缀）
    - http(s) URL（后端会下载转 base64）
    """
    from ..services.task_manager import store
    rec = store.records.get(record_id)
    if not rec:
        raise HTTPException(404, f"record {record_id} not found")

    avatar_data = body.avatar.strip()
    if not avatar_data:
        raise HTTPException(400, "avatar is empty")

    # 如果是 URL，下载转 base64
    if avatar_data.startswith(("http://", "https://")):
        import base64
        import httpx
        try:
            resp = httpx.get(avatar_data, timeout=15.0, follow_redirects=True)
            resp.raise_for_status()
            avatar_data = base64.b64encode(resp.content).decode("utf-8")
        except Exception as e:
            raise HTTPException(400, f"下载头像失败: {e}")

    # 如果带 data:image/... 前缀，去掉
    if avatar_data.startswith("data:"):
        avatar_data = avatar_data.split(",", 1)[-1]

    rec.avatar = avatar_data
    return {"ok": True, "record_id": record_id, "avatar": avatar_data[:80] + "..."}
