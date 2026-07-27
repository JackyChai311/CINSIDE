"""把 VerificationReport 导出为 Excel。"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..models import VerificationReport


def generate_excel(report: VerificationReport) -> bytes:
    """生成 .xlsx 文件字节流。"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Report")
    ws.title = "核验报告"

    # 标题
    ws.append(["CINSIDE 核验报告"])
    ws.append([])
    ws.append([
        "任务ID",
        "学生",
        "学校网址",
        "结论",
        "时间",
        "摘要",
    ])
    ws.append([
        report.task_id,
        report.record_name or report.record_id,
        report.university_url,
        report.overall,
        report.finished_at or report.started_at,
        report.summary or "",
    ])
    ws.append([])

    # 字段明细
    ws.append([
        "字段",
        "右侧选择器",
        "右侧标签",
        "左侧来源",
        "右侧值",
        "左侧值",
        "结果",
        "判定理由",
    ])

    status_fill = {
        "match": PatternFill("solid", fgColor="C6EFCE"),
        "mismatch": PatternFill("solid", fgColor="FFC7CE"),
        "missing": PatternFill("solid", fgColor="FFEB9C"),
        "error": PatternFill("solid", fgColor="B8CCE4"),
        "unknown": PatternFill("solid", fgColor="D9D9D9"),
    }

    for entry in report.entries:
        row = [
            entry.left_field,
            entry.right_selector,
            entry.right_label or "",
            entry.left_source,
            entry.right_value,
            entry.left_value,
            entry.match,
            entry.reasoning or "",
        ]
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=7)
        cell.fill = status_fill.get(entry.match, status_fill["unknown"])

    # 简单样式
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_filename(report: VerificationReport) -> str:
    """生成下载文件名。"""
    name = report.record_name or report.record_id
    ts = (report.finished_at or report.started_at or datetime.now().isoformat()).replace(":", "-")
    return f"verify_report_{name}_{ts}.xlsx"
