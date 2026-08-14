"""Excel 解析服务。

期望的 Excel 格式（首行表头，字段名可中可英，会做映射）：
  name / 姓名
  passport_no / 护照号
  nationality / 国籍
  birth_date / 出生日期
  gender / 性别
  passport_issue / 护照签发日期
  passport_expiry / 护照有效期
  email / 邮箱
  phone / 电话
  university_url / 大学申请页URL  (可选)
  university_name / 大学名称       (可选)

也支持 CSV（用 utf-8-sig 编码）。

字段识别策略：
  1. 优先用 AI（LLM）分析表头 + 样例数据，智能识别每列对应的标准字段
  2. AI 不可用时回退到硬编码别名表 _FIELD_ALIASES

多 Sheet 处理：
  - 读取 xlsx 的所有 sheet（包括隐藏 sheet、被分组折叠的 sheet）
  - 第一个非空 sheet 作为主表，生成 records
  - 其他 sheet 通过 name / passport_no / student_id 关联到主表，
    把额外字段合并到对应 record（冲突字段加 sheet 名前缀）
  - 关联不到的行：作为独立 record 追加
"""
from __future__ import annotations

import asyncio
import csv
import io
import os
from typing import Iterable

from ..models import ApplicantRecord, VERIFY_FIELDS
from .excel_ai_parser import ai_detect_column_mapping

# 字段别名映射：归一化成英文 key
_FIELD_ALIASES: dict[str, str] = {
    "name": "name", "姓名": "name", "fullname": "name", "applicant": "name",
    "student": "name", "studentname": "name", "studentname": "name", "full name": "name",
    "applicant name": "name", "chinese name": "name", "english name": "name",
    "大写": "name", "英文名": "name", "英文姓名": "name", "姓名大写": "name", "拼音": "name",
    "passport_no": "passport_no", "护照号": "passport_no", "护照": "passport_no",
    "passport_number": "passport_no",
    "passport no": "passport_no", "passport number": "passport_no", "passportno": "passport_no",
    "passport no.": "passport_no", "护照号码": "passport_no",
    "nationality": "nationality", "国籍": "nationality",
    "birth_date": "birth_date", "出生日期": "birth_date", "birthday": "birth_date", "dob": "birth_date",
    "birth date": "birth_date", "date of birth": "birth_date",
    "gender": "gender", "性别": "gender", "sex": "gender",
    "passport_issue": "passport_issue", "护照签发日期": "passport_issue", "issue_date": "passport_issue",
    "passport issue": "passport_issue", "date of issue": "passport_issue",
    "passport_expiry": "passport_expiry", "护照有效期": "passport_expiry", "expiry_date": "passport_expiry", "expire_date": "passport_expiry",
    "passport expiry": "passport_expiry", "date of expiry": "passport_expiry", "expiration date": "passport_expiry",
    "email": "email", "邮箱": "email", "e-mail": "email", "e mail": "email", "email address": "email",
    "phone": "phone", "电话": "phone", "mobile": "phone", "tel": "phone",
    "phone number": "phone", "mobile number": "phone", "联系电话": "phone",
    "university_url": "university_url", "大学申请页url": "university_url", "url": "university_url",
    "university_name": "university_name", "大学名称": "university_name", "university": "university_name",
    "school": "university_name", "school name": "university_name",
    "student_id": "student_id", "学号": "student_id", "student_no": "student_id", "sid": "student_id",
    "student id": "student_id", "student number": "student_id", "application id": "student_id",
    "application no": "student_id", "流水号": "student_id", "申请编号": "student_id",
}

# 用于跨 sheet 关联记录的"主键"字段（按优先级）
_JOIN_KEYS: tuple[str, ...] = (
    "passport_no", "护照号", "passport_number", "passport",
    "name", "姓名", "fullname", "applicant",
    "student_id", "学号", "student_no", "sid",
)


def _normalize_key(k: str) -> str:
    """归一化列名：去空格、去特殊字符、小写，再查别名表。"""
    import re
    k = (k or "").strip().lower()
    # 先查原始形式（可能直接命中，如 "Student Name"）
    if k in _FIELD_ALIASES:
        return _FIELD_ALIASES[k]
    # 去掉所有空格、下划线、连字符再查
    k_compact = re.sub(r"[\s_\-]+", "", k)
    if k_compact in _FIELD_ALIASES:
        return _FIELD_ALIASES[k_compact]
    # 带空格的版本再查一次（如 "full name" → "full name"）
    if k in _FIELD_ALIASES:
        return _FIELD_ALIASES[k]
    return k


def _safe_sheet_prefix(name: str) -> str:
    """把 sheet 名转成合法的字段前缀（snake_case）。"""
    import re
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", name.strip())
    s = re.sub(r"_+", "_", s).strip("_").lower()
    return s or "extra"


def _row_to_record(idx: int, row: dict[str, str], ai_mapping: dict[str, str] | None = None) -> tuple[ApplicantRecord, dict[str, str]]:
    """把一行数据转成 ApplicantRecord，同时返回 (record, detected_column_map)。
    detected_column_map: 原始列名 -> 标准字段名（如 {"大写": "name", "护照": "passport_no"}）
    """
    norm: dict[str, str] = {}
    col_map: dict[str, str] = {}
    for k, v in row.items():
        if v is None:
            continue
        # 优先用 AI 识别的列映射，其次用硬编码别名表
        if ai_mapping and k in ai_mapping:
            key = ai_mapping[k]
        else:
            key = _normalize_key(k)
        val = str(v).strip()
        if not val:
            continue
        if key in ("university_url", "university_name", "unknown", ""):
            continue
        # 1. 始终保留原始列名（用户在Excel中看到的表头）
        norm[k] = val
        # 2. 如果标准化 key 与原始列名不同，也存标准化 key（兼容现有代码直接访问 fields["name"] 等）
        if key != k:
            norm[key] = val
            # 记录自动识别的列映射（原始列 -> 标准字段）
            if key in VERIFY_FIELDS or key in ("student_id",):
                col_map[k] = key

    university_url = row.get("university_url") or row.get("大学申请页URL") or row.get("url") or ""
    university_name = row.get("university_name") or row.get("大学名称") or row.get("university") or ""

    return ApplicantRecord(
        record_id=f"rec-{idx:03d}",
        source="excel",
        fields=norm,
        university_url=university_url.strip() or None,
        university_name=university_name.strip() or None,
    ), col_map


def _read_xlsx_rows(source: str | bytes) -> tuple[list[str], list[list]]:
    """从路径或字节流读取 xlsx 的活动 sheet，返回 (header_keys, data_rows)。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("缺少依赖 openpyxl，请执行: pip install openpyxl") from e

    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    else:
        wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    header_keys = [str(h).strip() if h is not None else "" for h in header]
    data_rows: list[list] = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        data_rows.append(list(row))
    wb.close()
    return header_keys, data_rows


def _read_xlsx_all_sheets(source: str | bytes) -> list[tuple[str, list[str], list[list]]]:
    """读取 xlsx 的所有 sheet（包括隐藏 sheet），返回 [(sheet_name, header_keys, data_rows), ...]。

    - 不跳过被分组折叠（outline level）或隐藏的行/列：这些"折叠"的数据会被完整读出
    - 跳过完全空白的 sheet
    - 如果首行全部为空，自动生成列名（列1, 列2, ...），并将首行作为数据行
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("缺少依赖 openpyxl，请执行: pip install openpyxl") from e

    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    else:
        wb = load_workbook(source, read_only=True, data_only=True)

    sheets: list[tuple[str, list[str], list[list]]] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        header_keys = [str(h).strip() if h is not None else "" for h in header]
        data_rows: list[list] = []
        # 如果首行全部为空，说明没有表头：自动生成列名，首行作为数据
        if all(h == "" for h in header_keys):
            ncols = len(header_keys)
            header_keys = [f"列{i+1}" for i in range(ncols)]
            # 首行本身也作为数据行
            if not all(c is None or str(c).strip() == "" for c in header):
                data_rows.append(list(header))
        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            data_rows.append(list(row))
        if data_rows:
            sheets.append((ws.title, header_keys, data_rows))
    wb.close()
    return sheets


def _read_xls_all_sheets(source: str | bytes) -> list[tuple[str, list[str], list[list]]]:
    """读取旧版 .xls 的所有 sheet，返回 [(sheet_name, header_keys, data_rows), ...]。"""
    try:
        import xlrd
    except ImportError as e:
        raise RuntimeError("缺少依赖 xlrd，请执行: pip install xlrd") from e

    if isinstance(source, (bytes, bytearray)):
        wb = xlrd.open_workbook(file_contents=bytes(source), on_demand=True)
    else:
        wb = xlrd.open_workbook(source, on_demand=True)

    sheets: list[tuple[str, list[str], list[list]]] = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows < 1:
            continue
        header = ws.row_values(0)
        header_keys = [str(h).strip() if h is not None else "" for h in header]
        data_rows: list[list] = []
        # 如果首行全部为空，自动生成列名，首行作为数据
        if all(h == "" for h in header_keys):
            ncols = len(header_keys)
            header_keys = [f"列{i+1}" for i in range(ncols)]
            if not all(c is None or str(c).strip() == "" for c in header):
                data_rows.append(list(header))
        for r in range(1, ws.nrows):
            row = ws.row_values(r)
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            data_rows.append(list(row))
        if data_rows:
            sheets.append((sheet_name, header_keys, data_rows))
    wb.release_resources()
    return sheets


def _read_csv_rows(source: str | bytes) -> tuple[list[str], list[dict[str, str]]]:
    """从路径或字节流读取 csv，返回 (header_keys, row_dicts)。

    如果首行全部为空，自动生成列名（列1, 列2, ...）。
    """
    if isinstance(source, (bytes, bytearray)):
        text = source.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
    else:
        with open(source, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
    rows = list(reader)
    header_keys = list(rows[0].keys()) if rows else []
    # 如果所有列名为空，自动生成列名
    if header_keys and all(k is None or str(k).strip() == "" for k in header_keys):
        ncols = len(header_keys)
        new_keys = [f"列{i+1}" for i in range(ncols)]
        for row in rows:
            for old_k, new_k in zip(header_keys, new_keys):
                row[new_k] = row.pop(old_k, "")
        header_keys = new_keys
    return header_keys, rows


def _rows_to_records(
    header_keys: list[str],
    data_rows: list[list],
    ai_mapping: dict[str, str] | None = None,
) -> tuple[list[ApplicantRecord], dict[str, str]]:
    """把 xlsx 的 list-of-list 数据行转成 ApplicantRecord。
    返回 (records, detected_column_map)。
    """
    records: list[ApplicantRecord] = []
    detected_map: dict[str, str] = {}
    for i, row in enumerate(data_rows, start=1):
        d = {header_keys[k]: ("" if row[k] is None else str(row[k])) for k in range(len(header_keys)) if k < len(row)}
        rec, col_map = _row_to_record(i, d, ai_mapping)
        # 合并列映射（只取第一次出现的，避免后续行覆盖）
        for k, v in col_map.items():
            if k not in detected_map:
                detected_map[k] = v
        if rec.fields:
            records.append(rec)
    return records, detected_map


def _rowdicts_to_records(
    row_dicts: list[dict[str, str]],
    ai_mapping: dict[str, str] | None = None,
) -> tuple[list[ApplicantRecord], dict[str, str]]:
    """把 csv 的 list-of-dict 行转成 ApplicantRecord。
    返回 (records, detected_column_map)。
    """
    records: list[ApplicantRecord] = []
    detected_map: dict[str, str] = {}
    for i, row in enumerate(row_dicts, start=1):
        rec, col_map = _row_to_record(i, row, ai_mapping)
        for k, v in col_map.items():
            if k not in detected_map:
                detected_map[k] = v
        if rec.fields:
            records.append(rec)
    return records, detected_map


def _build_record_indexes(records: list[ApplicantRecord]) -> dict[str, dict[str, ApplicantRecord]]:
    """为主 records 建立 {归一化字段名: {值小写: record}} 的索引，用于跨 sheet 关联。"""
    indexes: dict[str, dict[str, ApplicantRecord]] = {}
    for r in records:
        for k_raw in _JOIN_KEYS:
            k_norm = _normalize_key(k_raw)
            val = (r.fields.get(k_norm) or r.fields.get(k_raw) or "").strip().lower()
            if not val:
                continue
            indexes.setdefault(k_norm, {})[val] = r
    return indexes


def _match_record(
    row_dict: dict[str, str],
    indexes: dict[str, dict[str, ApplicantRecord]],
) -> ApplicantRecord | None:
    """从 row_dict 找出能匹配的主 record。"""
    for k_raw, v in row_dict.items():
        if not v:
            continue
        k_norm = _normalize_key(k_raw)
        if k_norm not in indexes:
            continue
        val = v.strip().lower()
        if not val:
            continue
        hit = indexes[k_norm].get(val)
        if hit:
            return hit
    return None


def _merge_extra_row(
    target: ApplicantRecord,
    row_dict: dict[str, str],
    sheet_name: str,
    ai_mapping: dict[str, str] | None = None,
) -> int:
    """把附加 sheet 的 row_dict 字段合并到 target record，返回合并字段数。"""
    prefix = _safe_sheet_prefix(sheet_name)
    merged = 0
    for k, v in row_dict.items():
        if v is None:
            continue
        val = str(v).strip()
        if not val:
            continue
        # 优先用 AI 识别的列映射，其次用硬编码别名表
        if ai_mapping and k in ai_mapping:
            key = ai_mapping[k]
        else:
            key = _normalize_key(k)
        if key in ("university_url", "university_name", "unknown", ""):
            continue
        # 关联键本身不重复写入（用原始列名和标准化key都检查）
        if k in _JOIN_KEYS or key in _JOIN_KEYS or _normalize_key(k) in _JOIN_KEYS:
            continue
        # 原始列名（带prefix）作为存储key
        raw_prefixed_key = f"{prefix}_{k}"
        std_prefixed_key = f"{prefix}_{key}" if key != k else raw_prefixed_key
        # 主表已有该原始列名（带prefix）且非空：跳过
        if raw_prefixed_key in target.fields and target.fields[raw_prefixed_key]:
            continue
        target.fields[raw_prefixed_key] = val
        # 如果标准化key与原始列名不同，也存一份带prefix的标准化key
        if std_prefixed_key != raw_prefixed_key:
            target.fields[std_prefixed_key] = val
        merged += 1
    return merged


def parse_excel_or_csv(path: str) -> tuple[list[ApplicantRecord], dict[str, str]]:
    """根据后缀分流解析（同步版本，用硬编码别名表，不调用 AI）。
    返回 (records, detected_column_map)。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        header_keys, row_dicts = _read_csv_rows(path)
        return _rowdicts_to_records(row_dicts)
    if ext == ".xls":
        sheets = _read_xls_all_sheets(path)
    else:
        # xlsx：读取所有 sheet
        sheets = _read_xlsx_all_sheets(path)
    if not sheets:
        return [], {}
    main_name, main_header, main_rows = sheets[0]
    records, detected_map = _rows_to_records(main_header, main_rows)
    if len(sheets) > 1 and records:
        indexes = _build_record_indexes(records)
        for sheet_name, header, rows in sheets[1:]:
            for row in rows:
                d = {header[k]: ("" if row[k] is None else str(row[k]).strip()) for k in range(len(header)) if k < len(row)}
                matched = _match_record(d, indexes)
                if matched:
                    _merge_extra_row(matched, d, sheet_name)
                # 关联不上的行直接跳过（避免污染主表）
    return records, detected_map


async def parse_bytes(content: bytes, filename: str) -> tuple[list[ApplicantRecord], dict[str, str]]:
    """从前端上传的字节流解析（异步版本，优先用 AI 识别字段映射）。

    多 sheet 处理：
    - 主 sheet 调 AI 识别列映射
    - 附加 sheet 也调 AI 识别（如果配置了 LLM），否则用硬编码别名表
    - 通过 name / passport_no / student_id 把附加字段合并到主 record
    返回 (records, detected_column_map)，其中 detected_column_map 是原始列名 -> 标准字段名的映射。
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        header_keys, row_dicts = await asyncio.to_thread(_read_csv_rows, content)
        if not row_dicts:
            return [], {}
        # 准备样例数据给 AI（前 3 行）
        sample_rows = [list(r.values()) for r in row_dicts[:3]]
        ai_mapping = await ai_detect_column_mapping(header_keys, sample_rows)
        return _rowdicts_to_records(row_dicts, ai_mapping or None)

    # xls 或 xlsx：读取所有 sheet（包括隐藏的、被分组折叠的）
    # openpyxl/xlrd 读取是同步 CPU/IO 重活，放线程池避免阻塞事件循环
    if ext == ".xls":
        sheets = await asyncio.to_thread(_read_xls_all_sheets, content)
    else:
        sheets = await asyncio.to_thread(_read_xlsx_all_sheets, content)
    if not sheets:
        return [], {}

    # 第一个非空 sheet 作为主表
    main_name, main_header, main_rows = sheets[0]
    main_sample = main_rows[:3]
    main_ai_mapping = await ai_detect_column_mapping(main_header, main_sample)
    records, detected_map = _rows_to_records(main_header, main_rows, main_ai_mapping or None)

    # 其他 sheet 的字段合并到匹配的 record
    if len(sheets) > 1 and records:
        indexes = _build_record_indexes(records)
        for sheet_name, header, rows in sheets[1:]:
            # 附加 sheet 也调 AI 识别（如果配置了 LLM）
            extra_sample = rows[:3]
            extra_ai_mapping = await ai_detect_column_mapping(header, extra_sample)
            for row in rows:
                d = {header[k]: ("" if row[k] is None else str(row[k]).strip()) for k in range(len(header)) if k < len(row)}
                matched = _match_record(d, indexes)
                if matched:
                    _merge_extra_row(matched, d, sheet_name, extra_ai_mapping or None)
                # 关联不上的行：作为独立 record 追加（保留信息不丢失）
                else:
                    # 只在 row 有实质内容时才追加
                    non_empty = {k: v for k, v in d.items() if v and v.strip()}
                    if non_empty:
                        # 标记来源 sheet
                        non_empty["_source_sheet"] = sheet_name
                        rec, extra_map = _row_to_record(len(records) + 1, non_empty, extra_ai_mapping or None)
                        if rec.fields:
                            records.append(rec)
                            # 合并附加sheet的列映射
                            for k, v in extra_map.items():
                                if k not in detected_map:
                                    detected_map[k] = v

    return records, detected_map
