# -*- coding: utf-8 -*-
"""端到端验证导出高亮：模拟用户 Excel 结构（复合中文表头）+ LOOP 报告 highlights。"""
import sys, io
sys.path.insert(0, ".")

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from app.models import ApplicantRecord
from app.routers.upload import _apply_records_to_first_sheet

# 1. 构造与用户 Excel 相同结构的表头（复合中英文）
wb = Workbook()
ws = wb.active
ws.append(["申请编号", "姓 Family Name", "名 Given Name", "护照号码", "出生日期", "国籍", "LO"])
ws.append(["APP-001", "Syn-Khir", "Alisa", "670906781", "2008-07-17", "Russia", "RU"])
ws.append(["APP-002", "Ivanov", "Ivan", "67NO0906781", "2005-03-22", "Russia", "RU"])

# 2. 模拟 store.records（解析器会存原始 key + 规范 key）
recs = [
    ApplicantRecord(record_id="rec-001", source="excel", fields={
        "申请编号": "APP-001", "姓 Family Name": "Syn-Khir", "名 Given Name": "Alisa",
        "护照号码": "670906781", "出生日期": "2008-07-17", "国籍": "Russia", "LO": "RU",
        "surname": "Syn-Khir", "given_name": "Alisa", "passport_no": "670906781",
        "birth_date": "2008-07-17", "nationality": "Russia", "name": "Syn-Khir Alisa",
    }),
    ApplicantRecord(record_id="rec-002", source="excel", fields={
        "申请编号": "APP-002", "姓 Family Name": "Ivanov", "名 Given Name": "Ivan",
        "护照号码": "67NO0906781", "出生日期": "2005-03-22", "国籍": "Russia", "LO": "RU",
        "surname": "Ivanov", "given_name": "Ivan", "passport_no": "67NO0906781",
        "birth_date": "2005-03-22", "nationality": "Russia", "name": "Ivanov Ivan",
    }),
]

# 3. 模拟前端发来的 highlights（left_field 是规范字段名）
highlights = {
    "rec-001": {  # 第 1 行：用户已修复 → 黄色
        "surname": "fixed",
        "passport_no": "fixed",
    },
    "rec-002": {  # 第 2 行：未修复问题 → 红色
        "surname": "mismatch",
        "name": "mismatch",
        "passport_no": "mismatch",
        "birth_date": "missing",   # 琥珀
        "nationality": "mismatch",
    },
}

# 4. 执行写回+高亮
written = _apply_records_to_first_sheet(wb, recs, highlights)
print(f"written = {written}")

# 5. 读回每行的填色，验证
for row_idx in (2, 3):
    print(f"--- Row {row_idx} ---")
    for cell in wb.active[row_idx]:
        fill = cell.fill
        rgb = getattr(getattr(fill, "fgColor", None), "rgb", None) if fill and fill.patternType else None
        if rgb and isinstance(rgb, str) and rgb != "00000000":
            status = {"FFC7CE": "红", "FFEB9C": "琥珀", "C6EFCE": "绿", "FFFF00": "黄"}.get(rgb[-6:], rgb)
            print(f"  {wb.active.cell(row=1, column=cell.column).value}: {status}")
